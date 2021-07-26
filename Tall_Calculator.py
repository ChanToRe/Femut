#femur length를 이용하여 tall을 추정
#144번줄에서 저장위치 변경
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#Male
def M_pearson(number):
    result = 81.306 + 1.880 * number
    return result

def M_Trotter(number):
    result = 2.15 * number + 72.57
    return result

def M_huzii(number):
    result = (2.47 * (number*10) + 549.01)/10
    return result

#Female
def F_pearson(number):
    result = 72.844 + 1.945 * number
    return result

def F_huzii(number):
    result = (2.24 * (number*10) + 610.43)/10
    return result

wb = Workbook()
ws = wb.active
ws.title = 'result'

count = 2

#Sheet style of Excel
r1c1 = ws.cell(row = 1, column = 1)
r1c1.value = "Number"
r1c1.font = Font(bold=True)
r1c1.alignment = Alignment(horizontal='center')
r1c2 = ws.cell(row = 1, column = 2)
r1c2.value = "Sex"
r1c2.font = Font(bold=True)
r1c2.alignment = Alignment(horizontal='center')
r1c3 = ws.cell(row = 1, column = 3)
r1c3.value = "Femur Length"
r1c3.font = Font(bold=True)
r1c3.alignment = Alignment(horizontal='center')
r1c4 = ws.cell(row = 1, column = 4)
r1c4.value = "Pearson formula"
r1c4.font = Font(bold=True)
r1c4.alignment = Alignment(horizontal='center')
r1c5 = ws.cell(row = 1, column = 5)
r1c5.value = "Huzii formula"
r1c5.font = Font(bold=True)
r1c5.alignment = Alignment(horizontal='center')
r1c6 = ws.cell(row = 1, column = 6)
r1c6.value = "Trotter&Gleser formula"
r1c6.font = Font(bold=True)
r1c6.alignment = Alignment(horizontal='center')

#First Page
print("="*30)
print("="*9 + "femurtotall" + "="*10)
print("="*30)
print(" ")

while True : 
    print("Male : 1\nFemale : 2\nExit : 3")
    print(" ")
    Sex = input(">>> ")
    print(" ")
    if Sex == '1': #Male
        print("You chose a Male")
        print("You can use Pearson formula, Trotter&Glaser formula, Huzii formula")
        print("Enter to 0, you go to First Page")
        print(" ")
        while True:
            femur = float(input("Input the femur length >>> "))
            if femur != 0:
                pearson_result = M_pearson(femur)
                trotter_result = M_Trotter(femur)
                huzii_result = M_huzii(femur)
                
                #Sheet styles
                number = ws.cell(row = count, column = 1)
                number.value = count - 1
                number.alignment = Alignment(horizontal='center')
                Sexinsheet = ws.cell(row = count, column = 2)
                Sexinsheet.value = "Male"
                Sexinsheet.alignment = Alignment(horizontal='center')
                femur_length = ws.cell(row = count, column = 3)
                femur_length.value = femur
                femur_length.alignment = Alignment(horizontal='center')
                pearsoncell = ws.cell(row = count, column = 4)
                pearsoncell.alignment = Alignment(horizontal='center')
                huziicell = ws.cell(row = count, column = 5)
                huziicell.alignment = Alignment(horizontal='center')
                trottercell = ws.cell(row = count, column = 6)
                trottercell.alignment = Alignment(horizontal='center')

                #Output Result
                pearsoncell.value = pearson_result
                huziicell.value = huzii_result
                trottercell.value = trotter_result

                count += 1

                print("- Pearson formula : %0.3fcm" % pearson_result)
                print("- Huzii formula : %0.3fcm" % huzii_result)
                print("- Trotter&Glaser formula : %0.3fcm" % trotter_result)
                print(" ")

            elif femur == 0 :
                print(" ")
                print("=" * 30)
                print("=" * 2 + "Go to First Page" + "=" * 2)
                print("=" * 30)
                print(" ")
                break

    elif Sex == '2': #Female
        print("You chose a Female")
        print("You can use Pearson formula, Huzii formula")
        print("Enter to 0, you go to First Page")
        print(" ")
        while True:
            femur = float(input("Input the femur length >>> "))
            if femur != 0:
                pearson_result = F_pearson(femur)
                huzii_result = F_huzii(femur)
             
                number = ws.cell(row = count, column = 1)
                number.value = count - 1
                number.alignment = Alignment(horizontal='center')
                Sex = ws.cell(row = count, column = 2)
                Sex.value = "Female"
                Sex.alignment = Alignment(horizontal='center')
                femur_length = ws.cell(row = count, column = 3)
                femur_length.value = femur
                femur_length.alignment = Alignment(horizontal='center')
                pearsoncell = ws.cell(row = count, column = 4)
                pearsoncell.alignment = Alignment(horizontal='center')
                huziicell = ws.cell(row = count, column = 5)
                huziicell.alignment = Alignment(horizontal='center')

                #Output Result
                pearsoncell.value = pearson_result
                huziicell.value = huzii_result

                count += 1

                print("- Pearson formula : %0.3fcm"  % pearson_result)
                print("- Huzii formula : %0.3fcm" % huzii_result)
                print(" ")

            elif femur == 0 :
                print(" ")
                print("=" * 30)
                print("=" * 2 + "Go to First Page" + "=" * 2)
                print("=" * 30)
                print(" ")
                break
            
    elif Sex == '3': #Exit
        wb.save('/Users/jch/Desktop/result.xlsx')
        print("="*30)
        print("="*11 + "Exit" + "="*10)
        print("="*30)
        print(" ")
        break

    else: #Error
        print(" ")
        print("Error")
        print(" ")